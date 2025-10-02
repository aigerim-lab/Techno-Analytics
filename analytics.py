import os
from pathlib import Path
import pandas as pd
from sqlalchemy import create_engine
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from config import DB_URI

BASE = Path(__file__).resolve().parent
CHARTS = BASE / "charts"
EXPORTS = BASE / "exports"
SQL_FILE = BASE / "sql" / "new_queries.sql"

def ensure_dirs():
    CHARTS.mkdir(parents=True, exist_ok=True)
    EXPORTS.mkdir(parents=True, exist_ok=True)

def engine():
    return create_engine(DB_URI)

def load_queries(sql_path: Path) -> dict:
    """Читает файл SQL и собирает запросы по маркерам -- name: <id>.
       Если маркера нет — сгенерирует имя q01, q02, ..."""
    text = sql_path.read_text(encoding="utf-8")
    blocks = [b.strip() for b in text.split(";") if b.strip()]
    queries = {}
    auto_i = 1
    for block in blocks:
        lines = block.splitlines()
        qname = None
        body = []
        for ln in lines:
            if ln.lower().startswith("-- name:"):
                qname = ln.split(":", 1)[1].strip()
            else:
                body.append(ln)
        if not qname:
            qname = f"q{auto_i:02d}"
            auto_i += 1
        queries[qname] = "\n".join(body).strip()
    return queries

def save_fig(fig, path: Path, title: str, rows: int, what: str):
    fig.savefig(path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"[OK] Saved {path.name} | rows={rows} | {what}")

def pie_chart(df, labels_col, values_col, fname, title):
    fig, ax = plt.subplots()
    ax.pie(df[values_col], labels=df[labels_col], autopct="%1.1f%%", startangle=90)
    ax.axis("equal")
    ax.set_title(title)
    save_fig(fig, CHARTS / fname, title, len(df), f"pie: {title}")

def bar_chart(df, x, y, fname, title):
    fig, ax = plt.subplots()
    ax.bar(df[x], df[y])
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    ax.tick_params(axis='x', rotation=30)
    save_fig(fig, CHARTS / fname, title, len(df), f"bar: {title}")

def barh_chart(df, y, x, fname, title):
    fig, ax = plt.subplots()
    ax.barh(df[y], df[x])
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    save_fig(fig, CHARTS / fname, title, len(df), f"barh: {title}")

def line_chart(df, x, y, fname, title):
    fig, ax = plt.subplots()
    ax.plot(df[x], df[y], marker="o")
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    ax.tick_params(axis='x', rotation=30)
    save_fig(fig, CHARTS / fname, title, len(df), f"line: {title}")

def hist_chart(df, col, bins, fname, title):
    fig, ax = plt.subplots()
    ax.hist(df[col].dropna(), bins=bins)
    ax.set_title(title); ax.set_xlabel(col); ax.set_ylabel("count")
    save_fig(fig, CHARTS / fname, title, len(df), f"hist: {title}")

def scatter_chart(df, x, y, fname, title):
    fig, ax = plt.subplots()
    ax.scatter(df[x], df[y])
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    save_fig(fig, CHARTS / fname, title, len(df), f"scatter: {title}")

def export_to_excel(dfs: dict, filename: Path):
    filename = EXPORTS / filename
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
    # форматирование
    wb = load_workbook(filename)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        df = dfs[sheet.title]
        from pandas.api.types import is_numeric_dtype
        for idx, col in enumerate(df.columns, start=1):
            if is_numeric_dtype(df[col]):
                col_letter = get_column_letter(idx)
                last_row = len(df) + 1
                rng = f"{col_letter}2:{col_letter}{last_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                sheet.conditional_formatting.add(rng, rule)
    wb.save(filename)
    total_rows = sum(len(v) for v in dfs.values())
    print(f"[OK] Created {filename.name}, {len(dfs)} sheets, {total_rows} rows")

def main():
    ensure_dirs()
    e = engine()
    queries = load_queries(SQL_FILE)

    # 6 графиков
    plan = [
        ("q1_users_by_country", "pie",  dict(labels_col="country", values_col="user_count"),
         "01_pie_users_by_country.png", "Users by Country"),
        ("q2_events_by_genre", "bar",   dict(x="genre", y="event_count"),
         "02_bar_events_by_genre.png", "Events by Genre"),
        ("q3_top_artists_by_events", "barh", dict(y="artist", x="event_count"),
         "03_barh_top_artists.png", "Top Artists by Events"),
        ("q4_events_by_month", "line", dict(x="month", y="event_count"),
         "04_line_events_by_month.png", "Events per Month"),
        ("q5_user_age", "hist", dict(col="user_age", bins=10),
         "05_hist_user_age.png", "User Age Distribution"),
        ("q6_rating_vs_attendance", "scatter", dict(x="attended_events", y="avg_rating"),
         "06_scatter_rating_vs_attendance.png", "Avg Rating vs Attended Events"),
    ]

    dfs_for_excel = {}

    # строим 6 графиков
    for qname, ctype, params, outfile, title in plan:
        if qname not in queries:
            print(f"[WARN] {qname} not found in SQL file — skip")
            continue
        df = pd.read_sql(queries[qname], e)
        dfs_for_excel[qname] = df.copy()

        if ctype == "pie":
            pie_chart(df, fname=outfile, title=title, **params)
        elif ctype == "bar":
            bar_chart(df, fname=outfile, title=title, **params)
        elif ctype == "barh":
            barh_chart(df, fname=outfile, title=title, **params)
        elif ctype == "line":
            if "month" in df.columns:
                df["month"] = pd.to_datetime(df["month"])
            line_chart(df, fname=outfile, title=title, **params)
        elif ctype == "hist":
            hist_chart(df, fname=outfile, title=title, **params)
        elif ctype == "scatter":
            scatter_chart(df, fname=outfile, title=title, **params)

    
    for qname, sql in queries.items():
        try:
            df = pd.read_sql(sql, e)
            dfs_for_excel[qname] = df
            print(f"[Excel] Sheet {qname}: {len(df)} rows")
        except Exception as ex:
            print(f"[ERROR] query {qname}: {ex}")

    export_to_excel(dfs_for_excel, filename=Path("report_assignment2.xlsx"))

if __name__ == "__main__":
    main()
